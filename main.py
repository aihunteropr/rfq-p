# streamlit_app.py
import os, math, re, functools
from io import BytesIO
from collections import defaultdict

import streamlit as st
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from mistralai import Mistral
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ğŸ”‘  API & model
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
API_KEY = "mSohq6RlPC68LV1y0C27qs3gSfw4CbOe"
MODEL   = "mistral-large-latest"
client  = Mistral(api_key=API_KEY)

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ğŸ”§  helpers
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _clean(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    return str(x).strip()

def workbook_bytes(wb):
    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def wrap_all(ws):
    for r in ws.iter_rows():
        for c in r:
            c.alignment = Alignment(wrap_text=True,
                                    vertical="top", horizontal="left")

# --- Mistral wrappers ---------------------------------------------------------
@functools.lru_cache(maxsize=1024)
@retry(wait=wait_exponential(multiplier=1, min=2, max=20),
       stop=stop_after_attempt(5),
       retry=retry_if_exception_type(Exception))
def _fmt_uncached(txt):
    res = client.agents.complete(
        agent_id="ag:934c59a8:20250930:untitled-agent:03cdf810",
        messages=[{"role": "user", "content": txt}],
    )
    return re.sub(r"[`]+", "", res.choices[0].message.content)

def format_text(txt):
    try:  return _fmt_uncached(_clean(txt))
    except Exception as e:
        st.error(f"Format-text error â†’ {e}"); return _clean(txt)

@functools.lru_cache(maxsize=1024)
@retry(wait=wait_exponential(multiplier=1, min=2, max=20),
       stop=stop_after_attempt(5),
       retry=retry_if_exception_type(Exception))
def _manu_uncached(txt):
    res = client.chat.complete(
        model=MODEL,
        messages=[{"role":"user",
                   "content":("Extract manufacturer / maker names "
                              "separated by hyphen - plain list only\ncontent: "+txt)}])
    return res.choices[0].message.content

def manufacture_name(txt):
    try:  return _manu_uncached(_clean(txt))
    except Exception as e:
        st.error(f"Manufacturer-name error â†’ {e}"); return ""

# --- PDF helpers --------------------------------------------------------------
def pdf_text(pdf):
    return "".join(p.extract_text() for p in PdfReader(pdf).pages)

def pdf_clean_body(pdf):
    return re.sub(r"(REQUEST FOR QUOTATION[\s\S]*?RFQ Number \d+)",
                  "", pdf_text(pdf))

def parse_pdf(body, rfq_all):
    rfx = re.search(r"RFQ Number (\d+)", rfq_all)
    rfx_no = rfx.group(1) if rfx else "Unknown"

    pat_item = re.compile(
        r"(\d{5}) (\w?12\d{10}) (\d+(?:\.\d+)?)\s*(\w+) .*?(\d{2}\.\d{2}\.\d{4})",
        re.DOTALL)
    short_pat = re.compile(r"Short Text :(.*?)\n", re.DOTALL)
    po_pat    = re.compile(r"PO Material Text :(.*?)Agreement / LineNo.", re.DOTALL)

    items = pat_item.findall(body)
    short = short_pat.findall(body)
    po    = po_pat.findall(body)

    out=[]
    for i,it in enumerate(items):
        mat = it[1] if it[1].startswith(("B12","12","B16","15")) else ""
        out.append({"RFx Number":rfx_no,"RFx Item No":it[0],
                    "PR Item No":"","Material No":mat,
                    "Description":short[i] if i<len(short) else "",
                    "PO Text":po[i] if i<len(po) else "",
                    "QTY":it[2],"UOM":it[3]})
    return out


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ğŸ–¥ï¸  Streamlit UI
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
st.set_page_config(page_title="Data Processor", layout="wide",
                   initial_sidebar_state="collapsed")
st.markdown("""
<style>
.stButton button{background:#ff914d;color:#fff;border-radius:8px;
                 padding:10px 16px;margin-top:10px;}
.stExpander{background:#333;border-radius:10px;}
</style>
""", unsafe_allow_html=True)

col1,col2,col3=st.columns([2,2,1])

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# 1ï¸âƒ£  Excel Processor
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
with col1:
    st.subheader("ğŸ—ƒï¸ Excel Data Processor")
    techno = st.file_uploader("Techno-Commercial Envelope (.xls)",
                              type=["xls"], key="techno")

    with st.expander("Excel templates", True):
        upl_tpl = (st.file_uploader("Upload template (.xlsx)",
                                    type=["xlsx"], key="tpl_upl")
                   or "upload file - HTS.xlsx")
        fin_tpl = (st.file_uploader("Final Sheet template (.xlsx)",
                                    type=["xlsx"], key="tpl_fin")
                   or "FINAL SHEET.xlsx")

    suffix = st.text_input("Output name suffix", key="suffix_excel")

    if st.button("ğŸš€ Process Excel", key="btn_excel") and techno and suffix:
        try:
            rfx_no = re.search(r"\d+", techno.name).group()
            xls = pd.ExcelFile(techno)
            need = {'Description','InternalNote','Quantity','Unit of Measure'}
            sheet = next((s for s in xls.sheet_names
                          if need.issubset(set(pd.read_excel(xls,s,nrows=1).columns))),None)
            if not sheet: st.error("Required columns missing"); st.stop()

            df = pd.read_excel(techno, sheet_name=sheet, keep_default_na=False)
            valid = df[(df['Description'].str.strip().str.lower()!='item or lot description') &
                       df['Quantity'].astype(str).str.strip().ne('') &
                       df['Unit of Measure'].astype(str).str.strip().ne('') &
                       (df['Unit of Measure'].str.strip().str.lower()!='unit of measure')]

            # Upload workbook
            wb_u = load_workbook(upl_tpl); ws_u = wb_u.active
            for r in ws_u.iter_rows(min_row=2, max_row=ws_u.max_row):
                for c in r: c.value=None
            row,item=2,10
            for _,rec in valid.iterrows():
                ws_u[f"A{row}"]=rfx_no
                ws_u[f"B{row}"]=item
                ws_u[f"E{row}"]=rec['Description']
                ws_u[f"H{row}"]=rec['Unit of Measure']
                ws_u[f"G{row}"]=rec['Quantity']
                ws_u[f"F{row}"]=rec['InternalNote']
                ws_u[f"I{row}"]=rec.get('Number','')     # ARIBA Number
                item+=10; row+=1
            wrap_all(ws_u)

            # Final Sheet workbook
            wb_f = load_workbook(fin_tpl); ws_f = wb_f.active
            for r in ws_f.iter_rows(min_row=2, max_row=ws_f.max_row):
                for c in r: c.value=None
            row,item=2,10
            for _,rec in valid.iterrows():
                ws_f[f"A{row}"]=item
                ws_f[f"B{row}"]=rec['Description']
                ws_f[f"C{row}"]=rec['Quantity']
                ws_f[f"D{row}"]=rec['Unit of Measure']
                po = rec['InternalNote']
                ws_f[f"E{row}"]=format_text(po)
                ws_f[f"G{row}"]=manufacture_name(po)
                item+=10; row+=1
            wrap_all(ws_f)

            # Cache bytes in session_state
            st.session_state["excel_upload_bytes"] = workbook_bytes(wb_u)
            st.session_state["excel_final_bytes"]  = workbook_bytes(wb_f)
            st.session_state["excel_suffix"]       = suffix
            st.success("Excel processed âœ”ï¸")
        except Exception as e:
            st.error(f"âŒ {e}")

    # Download buttons persist
    if "excel_upload_bytes" in st.session_state:
        suf = st.session_state["excel_suffix"]
        st.download_button("ğŸ“¥ Download Upload file",
                           st.session_state["excel_upload_bytes"],
                           file_name=f"upload file - {suf}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_excel_up")
        st.download_button("ğŸ“¥ Download FINAL SHEET",
                           st.session_state["excel_final_bytes"],
                           file_name=f"FINAL SHEET - {suf}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_excel_fin")

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# 2ï¸âƒ£  PDF Processor
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
with col2:
    st.subheader("ğŸ“‘ PDF Data Processor")
    pdf = st.file_uploader("RFQ PDF", type=["pdf"], key="pdf")

    with st.expander("Excel templates", True):
        raw_tpl = (st.file_uploader("Raw template", type=["xlsx"], key="tpl_raw")
                   or "raw_template.xlsx")
        hts_tpl = (st.file_uploader("HTS template", type=["xlsx"], key="tpl_hts")
                   or "upload file - HTS.xlsx")
        fin_tpl_p = (st.file_uploader("Final Sheet template", type=["xlsx"],
                                      key="tpl_final_pdf")
                     or "FINAL SHEET.xlsx")

    hts_no = st.text_input("HTS number", key="hts_no")

    if st.button("ğŸš€ Process PDF", key="btn_pdf") and pdf and hts_no:
        try:
            data = parse_pdf(pdf_clean_body(pdf), pdf_text(pdf))

            # Upload workbook
            wb_up = load_workbook(hts_tpl); ws = wb_up.active
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for c in r: c.value=None
            row=2
            for rec in data:
                for col,letter in zip(
                    ["RFx Number","RFx Item No","PR Item No","Material No",
                     "Description","PO Text","QTY","UOM"],
                    list("ABCD")+list("EFGH")):
                    ws[f"{letter}{row}"]=rec[col]
                row+=1
            wrap_all(ws)

            # Final Sheet workbook
            wb_fin = load_workbook(fin_tpl_p); wsf = wb_fin.active
            for r in wsf.iter_rows(min_row=2, max_row=wsf.max_row):
                for c in r: c.value=None
            row=2
            for rec in data:
                wsf[f"A{row}"]=rec['RFx Item No']
                wsf[f"B{row}"]=rec['Description']
                wsf[f"C{row}"]=rec['QTY']
                wsf[f"D{row}"]=rec['UOM']
                wsf[f"E{row}"]=format_text(rec['PO Text'])
                wsf[f"G{row}"]=manufacture_name(rec['PO Text'])
                row+=1
            wrap_all(wsf)

            # Cache bytes
            st.session_state["pdf_upload_bytes"] = workbook_bytes(wb_up)
            st.session_state["pdf_final_bytes"]  = workbook_bytes(wb_fin)
            st.session_state["pdf_hts_no"]       = hts_no
            st.success("PDF processed âœ”ï¸")
        except Exception as e:
            st.error(f"âŒ {e}")

    if "pdf_upload_bytes" in st.session_state:
        no = st.session_state["pdf_hts_no"]
        st.download_button("ğŸ“¥ Download Upload file",
                           st.session_state["pdf_upload_bytes"],
                           file_name=f"upload file - {no}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_pdf_up")
        st.download_button("ğŸ“¥ Download FINAL SHEET",
                           st.session_state["pdf_final_bytes"],
                           file_name=f"FINAL SHEET - {no}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_pdf_fin")

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# 3ï¸âƒ£  HTS Cleaner & List Maker
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
with col3:
    st.subheader("ğŸ§¹ HTS Cleaner")
    hts_up = st.file_uploader("upload file â€“ HTS.xlsx",
                              type=["xlsx"], key="hts_clean")
    fin_tpl_opt = (st.file_uploader("Final Sheet template (opt)",
                                    type=["xlsx"], key="tpl_clean_fin")
                   or "FINAL SHEET.xlsx")

    if st.button("ğŸš€ Clean HTS", key="btn_clean_hts") and hts_up:
        try:
            wb_up = load_workbook(hts_up)
            wb_fin = load_workbook(fin_tpl_opt); wsf = wb_fin.active
            for r in wsf.iter_rows(min_row=2, max_row=wsf.max_row):
                for c in r: c.value=None
            up_ws = wb_up.active
            row=2
            for r in up_ws.iter_rows(min_row=2, max_row=up_ws.max_row):
                if not any(c.value for c in r): continue
                wsf[f"A{row}"]=r[1].value
                wsf[f"B{row}"]=r[4].value
                wsf[f"C{row}"]=r[6].value
                wsf[f"D{row}"]=r[7].value
                po = r[5].value or ""
                wsf[f"E{row}"]=format_text(po)
                wsf[f"G{row}"]=manufacture_name(po)
                row+=1
            wrap_all(wsf)
            st.session_state["clean_bytes"] = workbook_bytes(wb_fin)
            st.success("HTS cleaned âœ”ï¸")
        except Exception as e:
            st.error(f"âŒ {e}")

    if "clean_bytes" in st.session_state:
        st.download_button("ğŸ“¥ Download cleaned FINAL SHEET",
                           st.session_state["clean_bytes"],
                           file_name="FINAL SHEET - cleaned.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_clean")

    # â”€â”€ List Maker â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    st.subheader("ğŸ“ List Maker")
    final_xlsx = st.file_uploader("FINAL SHEET for manufacturers",
                                  type=["xlsx"], key="manuf")
    if st.button("ğŸš€ Build list", key="btn_list") and final_xlsx:
        try:
            df = pd.read_excel(final_xlsx)
            out = defaultdict(lambda:{"items":[],"emails":[]})
            for _,row in df.iterrows():
                mans=_clean(row.get("Manufacturer",""))
                if not mans: continue
                items=row["Line item number"]
                emails=[row[c] for c in df.columns
                        if ("mail" in c.lower() or "unnamed" in c.lower())
                        and pd.notna(row[c])]
                for m in [m.strip() for m in mans.split("-")]:
                    out[m]["items"].append(items)
                    out[m]["emails"].extend(emails)
            blob=[]
            for m,v in out.items():
                blob.append(f"Item {', '.join(map(str,sorted(set(v['items']))))}: {m}")
                if v["emails"]: blob.append("\n".join(v["emails"]))
                blob.append("")
            final="\n".join(blob)
            st.text_area("Output", final, height=300)
            from st_copy_to_clipboard import st_copy_to_clipboard
            st_copy_to_clipboard(final)
        except Exception as e:
            st.error(f"âŒ {e}")
